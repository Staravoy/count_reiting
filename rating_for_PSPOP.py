import openpyxl
from generate_name_colum import generate_name_column


def load_file_and_all_sheets():
    file = 'rozstanovka.xlsx'
    wb = openpyxl.load_workbook(file, data_only=True)
    list_all_sheets = wb.sheetnames
    return wb, list_all_sheets


def get_star_and_end_row(messange):
    while True:
        num_start_row = input(messange)
        if num_start_row.isnumeric():
            return num_start_row
        else:
            print("Це не число, спробуйте ще раз!")


def get_letter_work_column(messange):
    transform_letter_to_num = generate_name_column()
    while True:
        letter = input(messange)
        letter = letter.upper().replace(" ", "")
        if letter.isalpha() and letter.isascii():
            num_column = transform_letter_to_num[letter]
            return num_column
        else:
            print("Ви ввели не правильний символ, спробуйте ще раз!")


def get_all_names_force(star_row, end_row):
    work_file, work_sheets = load_file_and_all_sheets()
    sheet = work_file[work_sheets[0]]
    names_force = []
    for row in sheet.iter_rows(min_row=int(star_row), max_row=int(end_row), min_col=1, max_col=1):
        for cell in row:
            names_force.append(cell.value)
    return names_force


def create_new_excel_table():
    name = input("Введіть назву таблиці яка утвориться після збору інформації: ")
    new_table = openpyxl.Workbook()
    name_new_table = f'{name}.xlsx'
    new_table.save(name_new_table)
    return name_new_table


def fill_table(num_row, my_list):
    # Відкриваємо новий документ Excel
    wb = openpyxl.load_workbook('new_table.xlsx')
    # Вибираємо активний аркуш
    sheet = wb.active
    col = 1
    for x in my_list:
        sheet.cell(row=num_row, column=col, value=x)
        col += 1
    wb.save('new_table.xlsx')


# отримання даних необхідних для роботи
start = get_star_and_end_row("Введіть номер рядку ПЕРШОГО підрозділу в списку:")
end = get_star_and_end_row("Введіть номер рядку ОСТАННЬОГО підрозділу в списку:")
work_column = get_letter_work_column("Введіть літеру або літери, колонки яку треба опрацювати:")
name_new_table = create_new_excel_table()
print("Збір даних розпочато...")

# отримуємо імена всіх підрозділів
name_forces = get_all_names_force(start, end)
# завантажили робочу книгу з усіма листами
file, sheets = load_file_and_all_sheets()


for name in range(0, len(name_forces)):
    list_for_new_table = []
    list_for_new_table.append(name_forces[name])
    for one_sheet in sheets:
        sheet = file[one_sheet]
        # шукаємо номер рядку підрозділу на кожному аркуші
        for i in range(int(start), int(end) + 1):
            cell_value = sheet.cell(row=i, column=1).value
            if cell_value == name_forces[name]:
                row_one_force = i
        cell = sheet.cell(row=row_one_force, column=work_column).value
        if not cell:
            continue
        list_for_new_table.append(cell)
    print(list_for_new_table)

