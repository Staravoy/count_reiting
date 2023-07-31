
import os
import openpyxl
import sys
from generate_name_colum import generate_name_column
import tkinter as tk
from tkinter import filedialog

def get_abs_path(file_name):
    # Отримати абсолютний шлях до файлу, де знаходиться виконуваний файл
    current_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    abs_path = os.path.join(current_dir, file_name)
    return abs_path


def dialog_window():
    # Створення головного вікна
    root = tk.Tk()
    root.withdraw()
    # Відкриваємо діалогове вікно для вибору файлу
    file_path = filedialog.askopenfilename(title="Оберіть файл для обробки", filetypes=[("All Files", "*.*")])
    return file_path

def load_file_and_all_sheets():
    file_path = dialog_window()
    wb = openpyxl.load_workbook(get_abs_path(file_path), data_only=True)
    list_all_sheets = wb.sheetnames
    wb.close()
    return wb, list_all_sheets


def get_star_and_end_row(messange):
    while True:
        num_start_row = input(messange)
        if num_start_row.isnumeric():
            return num_start_row
        else:
            print("Це не число, спробуйте ще раз!")


def get_letter_work_column(messange):
    transform_letter_to_num = generate_name_column()
    while True:
        letter = input(messange)
        letter = letter.upper().replace(" ", "")
        if letter.isalpha() and letter.isascii():
            num_column = transform_letter_to_num[letter]
            return num_column
        else:
            print("Ви ввели не правильний символ, спробуйте ще раз!")


def get_all_names_force(star_row, end_row, work_file, work_sheets):
    sheet = work_file[work_sheets[0]]
    names_force = []
    for row in sheet.iter_rows(min_row=int(star_row), max_row=int(end_row), min_col=1, max_col=1):
        for cell in row:
            names_force.append(cell.value)
    work_file.close()
    return names_force


def create_new_excel_table():
    name = input("Введіть назву таблиці яка утвориться після збору інформації: ")
    new_table = openpyxl.Workbook()
    name_new_table = f'{name}.xlsx'
    new_table.save(get_abs_path(name_new_table))
    new_table.close()
    return name_new_table


def fill_table(name_file, num_row, my_list):
    # Відкриваємо новий документ Excel
    name = name_file
    wb = openpyxl.load_workbook(get_abs_path(name))
    # Вибираємо активний аркуш
    sheet = wb.active
    col = 1
    for x in my_list:
        sheet.cell(row=num_row, column=col, value=x)
        col += 1
    wb.save(get_abs_path(name))
    wb.close()

while True:
    # отримання даних необхідних для роботи
    print("Слава Україні!!!")
    print("Вас вітає програма: ПІДРАХУЙКА!!!")
    print("її завдання полегшити процес збирання інформації.")
    input("Ну, що до роботи? Натисніть ENTER")
    print("Зчитуємо файл...")
    # завантажили робочу книгу з усіма листами
    file, sheets = load_file_and_all_sheets()
    start = get_star_and_end_row("Введіть номер рядку ПЕРШОГО підрозділу в списку:")
    end = get_star_and_end_row("Введіть номер рядку ОСТАННЬОГО підрозділу в списку:")
    work_column = get_letter_work_column("Введіть літеру або літери, колонки яку треба опрацювати:")
    # отримуємо назву нового документу та створюємо його
    name_new_table = create_new_excel_table()
    print("Збір даних розпочато...")

    # отримуємо імена всіх підрозділів
    name_forces = get_all_names_force(start, end, file, sheets)

    # створюємо шапку таблиці з обрізаних назв листів
    list_sheet_name = [" "]
    for sheet in sheets:
        list_sheet_name.append(sheet[:-5])
    fill_table(name_new_table, 1, list_sheet_name)


    num_row = 2
    for name in range(0, len(name_forces)):
        list_for_new_table = []
        list_for_new_table.append(name_forces[name])
        for one_sheet in sheets:
            sheet = file[one_sheet]
            # шукаємо номер рядку підрозділу на кожному аркуші
            for i in range(int(start), int(end) + 1):
                cell_value = sheet.cell(row=i, column=1).value
                if cell_value == name_forces[name]:
                    row_one_force = i
            cell = sheet.cell(row=row_one_force, column=work_column).value
            if not cell:
                continue
            list_for_new_table.append(cell)

        fill_table(name_new_table, num_row, list_for_new_table)
        num_row += 1

    print("Збір даних завершено!")
    question = input("Чи бажаєте продовжити роботу? (так/ні)")
    if question == 'ні':
        break


