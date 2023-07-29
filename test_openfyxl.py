import openpyxl
from generate_name_colum import generate_name_column

file = 'rozstanovka.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
all_sheets = wb.sheetnames

# оголошую словник
dictonary_for_table = {}

# отримуємо усі назви підрозділів
start_row = input("Start row = ")
start_row = int(start_row)
end_row = input("End row = ")
end_row = int(end_row)
sheet = wb[all_sheets[0]]
names_forces = []
for i in range(start_row, end_row+1):
    forces = sheet.cell(row=i, column=1).value
    if not forces:
        continue
    names_forces.append(forces)

# отримуємо літеру стовбчика для обробки
def num_column():
    get_letter_col = input("Enter letter column for processing: ")
    get_letter_col = get_letter_col.upper()
    return get_letter_col

# отримання списку відповідності літер до номерів
letter_to_num_column = generate_name_column()

work_num_column = num_column()

while not work_num_column.isalpha():
    print("This is not a letter.")
    work_num_column = num_column()
work_num_column = letter_to_num_column[work_num_column]


# перебираємо кожну сторінку і отримуємо номер рядку
value = []
for one_sheet in all_sheets:
    sheet = wb[one_sheet]
    # шукаємо номер рядку підрозділу на кожнму аркуші
    for i in range(start_row, end_row + 1):
        cell_value = sheet.cell(row=i, column=1).value
        if cell_value == names_forces[-1]:
            row_one_force = i
    cell = sheet.cell(row=row_one_force, column=work_num_column).value
    if not cell:
        continue
    value.append(cell)
# формування словнику
dictonary_for_table[names_forces[-1]]=value
print(dictonary_for_table)


