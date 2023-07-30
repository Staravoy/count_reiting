import openpyxl

def load_file_and_all_sheets():
    file = 'rozstanovka.xlsx'
    wb = openpyxl.load_workbook(file, data_only=True)
    list_all_sheets = wb.sheetnames
    return wb, list_all_sheets




work_file, work_sheets = load_file_and_all_sheets()
sheet = work_file[work_sheets[0]]
names_force = []
num_row = 7
for row in sheet.iter_rows(min_row=7, max_row=26, min_col=1, max_col=1):
    for cell in row:
        print(f"Значення: {cell.value} - номер рядку {num_row}")
        num_row += 1
print(names_force)